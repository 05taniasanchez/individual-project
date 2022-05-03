from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
root = Tk()
root.title("CSUMB Student App")
root.geometry('430x300')
root.resizable(0,0)

img = PhotoImage(file="csumbtransparent.png")
label =Label(root, image=img)
label.place(x=100, y=40)

wb = Workbook()
wb = load_workbook("studentname.xlsx")
ws = wb.active

column_1 = ws['A']


def get_name():
    list=''
    for cell in column_1:
        list = f'{list + str(cell.value)}\n'

        label.config(text=list)

button = Button(root, text="Student's Name",fg="black", bg= "light blue", command=get_name).place(x=150, y=120)


label = Label(root, fg="black", bg= "light blue", text="")
label.place(x=150,y =145)

root.mainloop()